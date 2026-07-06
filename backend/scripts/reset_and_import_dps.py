"""
Reset Daffodils Public School (DPS) from mock seed data to the real
Nursery-Class 8 roster, and roll the tenant onto a fresh 2026-2027 academic year.

Admission-number convention (approved by school owner):
  DPSN-2026-{roll:03d}   Nursery
  DPSK1-2026-{roll:03d}  K.G. I
  DPSK2-2026-{roll:03d}  K.G. II
  DPS{n}-2026-{roll:03d} Class 1..8

Class 6 has no real student data in the source file (all 19 rows are blank
roll-number placeholders, no name/DOB/gender/phone) — its class/section shell
is created with zero students; the real roster must be imported later.

Missing DOB -> 2000-01-01, missing phone -> 0000000000 (matches the PMIC
import convention in import_pmic_science_commerce.py). All real rows already
have gender recorded (M/F), no placeholder needed there.
"""

import asyncio
import io
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import asyncpg
import openpyxl
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent))
from services.student import import_students

TENANT_SLUG = "daffodilspublicschool"
SOURCE_XLSX = Path(__file__).parent.parent.parent / "School_docs" / "Daffodils" / "STUDENTS  INFO. $.xlsx"

NEW_AY_NAME = "2026-2027"
NEW_AY_START = date(2026, 4, 1)
NEW_AY_END = date(2027, 3, 31)
ADM_YEAR = "2026"

PLACEHOLDER_DOB = "2000-01-01"
PLACEHOLDER_PHONE = "0000000000"

# source CLASS cell -> (db class name, admission prefix, numeric_order, import real rows?)
CLASS_DEFS = {
    "NUR":    ("Nursery",  "DPSN",  1, True),
    "K.G.I":  ("K.G. I",   "DPSK1", 2, True),
    "K.G.II": ("K.G. II",  "DPSK2", 3, True),
    1: ("Class 1", "DPS1", 4, True),
    2: ("Class 2", "DPS2", 5, True),
    3: ("Class 3", "DPS3", 6, True),
    4: ("Class 4", "DPS4", 7, True),
    5: ("Class 5", "DPS5", 8, True),
    6: ("Class 6", "DPS6", 9, False),   # no real data in source -- shell only
    7: ("Class 7", "DPS7", 10, True),
    8: ("Class 8", "DPS8", 11, True),
}

TEMPLATE_HEADERS = [
    "Admission No", "First Name", "Last Name", "Class", "Section",
    "Roll No", "Date of Birth", "Gender", "Parent Phone", "Hosteler", "Transport",
]

# One-off fix for a truncated year the general cleaner below can't recover
# (source row 262: "05.04.208" for K.G.II roll 16 "SMILY KUMARI"; neighbouring
# K.G.II rows are all born 2017-2019, so 2018 is the intended year).
DOB_OVERRIDES = {
    ("K.G.II", 16): "2018-04-05",
}


def normalise_phone(raw) -> str:
    digits = re.sub(r"\D", "", str(raw) if raw is not None else "")
    if len(digits) >= 10:
        return digits[-10:]
    return PLACEHOLDER_PHONE


def parse_dob(cell, class_label, roll) -> str:
    override = DOB_OVERRIDES.get((class_label, roll))
    if override:
        return override
    if cell is None:
        return PLACEHOLDER_DOB
    if isinstance(cell, (datetime, date)):
        d = cell.date() if isinstance(cell, datetime) else cell
        return d.strftime("%Y-%m-%d")
    # Source has a handful of typos: letter O for zero, stray spaces, doubled
    # separators (e.g. "20.08.2O13", "3. 1.2013", "19..11.2018").
    s = str(cell).strip().upper().replace("O", "0")
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[.\-/]+", ".", s)
    try:
        return datetime.strptime(s, "%d.%m.%Y").strftime("%Y-%m-%d")
    except ValueError:
        return PLACEHOLDER_DOB


def read_source() -> list[dict]:
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None and str(c).strip() for c in r)]
    wb.close()

    records = []
    skipped_no_data = 0
    for row in rows[1:]:
        cls_raw = row[3]
        if cls_raw not in CLASS_DEFS:
            continue
        db_name, prefix, _order, do_import = CLASS_DEFS[cls_raw]
        if not do_import:
            continue

        first = str(row[1]).strip() if row[1] else ""
        last = str(row[2]).strip() if row[2] else ""
        if not first:
            skipped_no_data += 1
            continue

        roll = int(str(row[5]).strip())
        gender_raw = str(row[6]).strip().upper() if row[6] else ""
        gender = {"M": "Male", "F": "Female", "MALE": "Male", "FEMALE": "Female"}.get(gender_raw, "Other")
        dob = parse_dob(row[7], cls_raw, roll)
        phone = normalise_phone(row[10])
        transport = "Yes" if row[11] and str(row[11]).strip().upper() == "YES" else "No"
        hosteler = "Yes" if row[12] and str(row[12]).strip().upper() == "YES" else "No"

        records.append({
            "Admission No":  f"{prefix}-{ADM_YEAR}-{roll:03d}",
            "First Name":    first,
            "Last Name":     last,
            "Class":         db_name,
            "Section":       "A",
            "Roll No":       str(roll),
            "Date of Birth": dob,
            "Gender":        gender,
            "Parent Phone":  phone,
            "Hosteler":      hosteler,
            "Transport":     transport,
        })

    print(f"Parsed {len(records)} real student rows (skipped {skipped_no_data} blank-name rows)")
    return records


def build_template_xlsx(records: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(TEMPLATE_HEADERS)
    for rec in records:
        ws.append([rec[h] for h in TEMPLATE_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def reset_tenant(conn: asyncpg.Connection, tenant_id) -> dict:
    """Delete mock seed students/classes and roll onto a fresh academic year."""
    async with conn.transaction():
        student_ids = [r["id"] for r in await conn.fetch(
            "SELECT id FROM students WHERE tenant_id = $1", tenant_id)]
        class_ids = [r["id"] for r in await conn.fetch(
            "SELECT id FROM classes WHERE tenant_id = $1", tenant_id)]

        if student_ids:
            await conn.execute("DELETE FROM mark_entries WHERE student_id = ANY($1::uuid[])", student_ids)
            await conn.execute("DELETE FROM exam_component_marks WHERE student_id = ANY($1::uuid[])", student_ids)
            await conn.execute(
                """DELETE FROM fee_payment_items
                   WHERE ledger_id IN (SELECT id FROM fee_ledger WHERE student_id = ANY($1::uuid[]))
                      OR payment_id IN (SELECT id FROM fee_payments WHERE student_id = ANY($1::uuid[]))""",
                student_ids,
            )
            await conn.execute("DELETE FROM fee_ledger WHERE student_id = ANY($1::uuid[])", student_ids)
            await conn.execute("DELETE FROM fee_payments WHERE student_id = ANY($1::uuid[])", student_ids)
            await conn.execute("DELETE FROM admissions WHERE student_id = ANY($1::uuid[])", student_ids)
            await conn.execute("DELETE FROM students WHERE tenant_id = $1", tenant_id)

        if class_ids:
            await conn.execute("DELETE FROM attendance_sessions WHERE class_id = ANY($1::uuid[])", class_ids)
            await conn.execute("DELETE FROM exam_subjects WHERE class_id = ANY($1::uuid[])", class_ids)
            await conn.execute("DELETE FROM homework_posts WHERE class_id = ANY($1::uuid[])", class_ids)
            await conn.execute("DELETE FROM staff_class_assignments WHERE class_id = ANY($1::uuid[])", class_ids)
            await conn.execute("DELETE FROM timetable_slots WHERE class_id = ANY($1::uuid[])", class_ids)
            await conn.execute(
                "UPDATE admissions SET applying_class_id = NULL WHERE applying_class_id = ANY($1::uuid[])",
                class_ids,
            )
            await conn.execute("DELETE FROM classes WHERE tenant_id = $1", tenant_id)

        await conn.execute(
            "UPDATE academic_years SET is_current = FALSE, status = 'archived' "
            "WHERE tenant_id = $1 AND is_current = TRUE",
            tenant_id,
        )
        new_ay = await conn.fetchrow(
            """INSERT INTO academic_years (tenant_id, name, start_date, end_date, is_current, status)
               VALUES ($1, $2, $3, $4, TRUE, 'active') RETURNING id""",
            tenant_id, NEW_AY_NAME, NEW_AY_START, NEW_AY_END,
        )
        new_ay_id = new_ay["id"]

        seen_names = set()
        for db_name, _prefix, order, _do_import in CLASS_DEFS.values():
            if db_name in seen_names:
                continue
            seen_names.add(db_name)
            row = await conn.fetchrow(
                "INSERT INTO classes (tenant_id, name, numeric_order) VALUES ($1, $2, $3) RETURNING id",
                tenant_id, db_name, order,
            )
            await conn.execute(
                "INSERT INTO sections (tenant_id, class_id, name) VALUES ($1, $2, 'A')",
                tenant_id, row["id"],
            )

    return {
        "deleted_students": len(student_ids),
        "deleted_classes": len(class_ids),
        "new_academic_year_id": new_ay_id,
    }


async def main():
    database_url = os.environ["DATABASE_URL"]
    conn = await asyncpg.connect(database_url)

    tenant = await conn.fetchrow("SELECT id FROM tenants WHERE slug = $1", TENANT_SLUG)
    if not tenant:
        print(f"ERROR: tenant '{TENANT_SLUG}' not found"); await conn.close(); return
    tenant_id = tenant["id"]

    records = read_source()
    if not records:
        print("Nothing to import."); await conn.close(); return

    reset_result = await reset_tenant(conn, tenant_id)
    print(f"Deleted {reset_result['deleted_students']} mock students, "
          f"{reset_result['deleted_classes']} mock classes")
    print(f"New academic year: {NEW_AY_NAME} ({reset_result['new_academic_year_id']})")

    xlsx_bytes = build_template_xlsx(records)
    result = await import_students(conn, tenant_id, reset_result["new_academic_year_id"], xlsx_bytes)
    await conn.close()

    print("\n=== Import result ===")
    print(f"  created : {result['created']}")
    print(f"  updated : {result['updated']}")
    if result["errors"]:
        print(f"  errors  : {len(result['errors'])}")
        for e in result["errors"]:
            print(f"    {e}")
    else:
        print("  errors  : 0")

    if result["errors"]:
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
