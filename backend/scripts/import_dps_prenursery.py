"""
Add Pre-Nursery to Daffodils Public School (DPS) and import its roster.

Follow-up to reset_and_import_dps.py: the source file has a "PRE. NUR" class
that predates Nursery in the school's actual class ladder but wasn't in the
original CLASS_DEFS (Nursery was treated as the youngest class at the time).
This script creates the Pre-Nursery class/section shell (numeric_order 0, so
it sorts before Nursery) and imports its 30-student roster into the existing
tenant + current academic year — it does NOT reset/delete anything.

Admission-number convention (matches reset_and_import_dps.py): DPSPN-2026-{roll:03d}

Data-quality note: source row for roll 30 (RICHA MAHTO) has shifted columns
(gender cell contains a date string, DOB cell is empty, phone is 9 digits)
— handled with the same placeholder conventions already used for the
original 352-student import (unrecognised gender -> "Other", missing DOB ->
2000-01-01, <10-digit phone -> 0000000000), not a new judgment call.
"""

import asyncio
import io
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import asyncpg
import openpyxl
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent))
from services.student import import_students

TENANT_SLUG = "daffodilspublicschool"
SOURCE_XLSX = Path(__file__).parent.parent.parent / "School_docs" / "Daffodils" / "STUDENTS  INFO. $.xlsx"

ADM_YEAR = "2026"
CLASS_LABEL = "Pre-Nursery"
SOURCE_CLASS_VALUE = "PRE. NUR"
PREFIX = "DPSPN"
NUMERIC_ORDER = 0

PLACEHOLDER_DOB = "2000-01-01"
PLACEHOLDER_PHONE = "0000000000"

TEMPLATE_HEADERS = [
    "Admission No", "First Name", "Last Name", "Class", "Section",
    "Roll No", "Date of Birth", "Gender", "Parent Phone", "Hosteler", "Transport",
]


def normalise_phone(raw) -> str:
    digits = re.sub(r"\D", "", str(raw) if raw is not None else "")
    if len(digits) >= 10:
        return digits[-10:]
    return PLACEHOLDER_PHONE


def parse_dob(cell) -> str:
    if cell is None:
        return PLACEHOLDER_DOB
    if isinstance(cell, (datetime, date)):
        d = cell.date() if isinstance(cell, datetime) else cell
        return d.strftime("%Y-%m-%d")
    s = str(cell).strip().upper().replace("O", "0")
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[.\-/]+", ".", s)
    try:
        return datetime.strptime(s, "%d.%m.%Y").strftime("%Y-%m-%d")
    except ValueError:
        return PLACEHOLDER_DOB


def read_source() -> list[dict]:
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None and str(c).strip() for c in r)]
    wb.close()

    records = []
    skipped_no_data = 0
    for row in rows[1:]:
        if row[3] != SOURCE_CLASS_VALUE:
            continue

        first = str(row[1]).strip() if row[1] else ""
        last = str(row[2]).strip() if row[2] else ""
        if not first:
            skipped_no_data += 1
            continue

        roll = int(str(row[5]).strip())
        gender_raw = str(row[6]).strip().upper() if row[6] else ""
        gender = {"M": "Male", "F": "Female", "MALE": "Male", "FEMALE": "Female"}.get(gender_raw, "Other")
        dob = parse_dob(row[7])
        phone = normalise_phone(row[10])
        transport = "Yes" if row[11] and str(row[11]).strip().upper() == "YES" else "No"
        hosteler = "Yes" if row[12] and str(row[12]).strip().upper() == "YES" else "No"

        records.append({
            "Admission No":  f"{PREFIX}-{ADM_YEAR}-{roll:03d}",
            "First Name":    first,
            "Last Name":     last,
            "Class":         CLASS_LABEL,
            "Section":       "A",
            "Roll No":       str(roll),
            "Date of Birth": dob,
            "Gender":        gender,
            "Parent Phone":  phone,
            "Hosteler":      hosteler,
            "Transport":     transport,
        })

    print(f"Parsed {len(records)} Pre-Nursery rows (skipped {skipped_no_data} blank-name rows)")
    return records


def build_template_xlsx(records: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(TEMPLATE_HEADERS)
    for rec in records:
        ws.append([rec[h] for h in TEMPLATE_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def ensure_class(conn: asyncpg.Connection, tenant_id) -> None:
    exists = await conn.fetchval(
        "SELECT 1 FROM classes WHERE tenant_id = $1 AND name = $2", tenant_id, CLASS_LABEL
    )
    if exists:
        return
    async with conn.transaction():
        row = await conn.fetchrow(
            "INSERT INTO classes (tenant_id, name, numeric_order) VALUES ($1, $2, $3) RETURNING id",
            tenant_id, CLASS_LABEL, NUMERIC_ORDER,
        )
        await conn.execute(
            "INSERT INTO sections (tenant_id, class_id, name) VALUES ($1, $2, 'A')",
            tenant_id, row["id"],
        )
    print(f"Created class '{CLASS_LABEL}' (numeric_order={NUMERIC_ORDER}) + section A")


async def main():
    database_url = os.environ["DATABASE_URL"]
    conn = await asyncpg.connect(database_url)

    tenant = await conn.fetchrow("SELECT id FROM tenants WHERE slug = $1", TENANT_SLUG)
    if not tenant:
        print(f"ERROR: tenant '{TENANT_SLUG}' not found"); await conn.close(); return
    tenant_id = tenant["id"]

    ay = await conn.fetchrow(
        "SELECT id FROM academic_years WHERE tenant_id = $1 AND is_current = TRUE", tenant_id
    )
    if not ay:
        print("ERROR: no current academic year found"); await conn.close(); return

    records = read_source()
    if not records:
        print("Nothing to import."); await conn.close(); return

    await ensure_class(conn, tenant_id)

    xlsx_bytes = build_template_xlsx(records)
    result = await import_students(conn, tenant_id, ay["id"], xlsx_bytes)
    await conn.close()

    print("\n=== Import result ===")
    print(f"  created : {result['created']}")
    print(f"  updated : {result['updated']}")
    if result["errors"]:
        print(f"  errors  : {len(result['errors'])}")
        for e in result["errors"]:
            print(f"    {e}")
    else:
        print("  errors  : 0")

    if result["errors"]:
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
