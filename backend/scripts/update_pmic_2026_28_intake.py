"""
Update Class 11 (2026-28 intake) students for Premchand Mahto Inter College
(PMIC) -- Arts, Science, and Commerce -- from freshly re-supplied source
files. Owner: "we are just updating new students that exist in the files
i sent you, we already have students running."

Same pattern as import_pmic_science_commerce.py (Science/Commerce, 2025-27 +
2026-28), extended to also cover Arts using the same admission-number
scheme already live in prod (A-2026-{roll:03d}, confirmed against existing
Class 11 Arts rows before writing this script). Only the 2026-28 (Class 11)
files were re-supplied this round -- the 2025-27 "2nd year" Class 12 files
are untouched, so this script does not touch Class 12.

Admission-ID convention (matches existing prod data):
  A-2026-{roll:03d}  →  Class 11, Arts
  S-2026-{roll:03d}  →  Class 11, Science
  C-2026-{roll:03d}  →  Class 11, Commerce

import_students is an admission_no-keyed upsert (services/student.py), so
re-running this is safe: existing rows get corrected fields (e.g. a phone
number fixed in the new source), new rows (the roster grew: Arts 90→107,
Science 15→25, Commerce 6→7 rows in the source files) get created. Nothing
is deleted.

Name split: first whitespace token → First Name, remainder → Last Name.
Missing DOB   → 2000-01-01 (placeholder, matches existing practice).
Missing phone → 0000000000 (matches existing practice).
Gender        → Other (source sheets have no gender column; matches
                 existing practice for all PMIC students so far).
Caste / Aadhar / Guardian Aadhar / Passing Board / Passing % columns exist
in the source sheets but are NOT captured -- students has no columns for
them (out of scope for this pass; flag to owner if these need to be
modeled later).
"""

import asyncio
import io
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import asyncpg
import openpyxl
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent))
from services.student import import_students

TENANT_SLUG = "premchandmahtoic"

BASE = Path(__file__).parent.parent.parent / "School_docs" / "Premchand Mahto Inter College"

# (source_file, admission_prefix, adm_year, class_name, section_name)
SOURCES = [
    ("Arts/session 2026-28 Arts.xlsx",           "A", "2026", "Class 11", "Arts"),
    ("Science/session 2026-28 science.xlsx",     "S", "2026", "Class 11", "Science"),
    ("Commerce/session 2026-28 - commerce.xlsx", "C", "2026", "Class 11", "Commerce"),
]

PLACEHOLDER_DOB = "2000-01-01"
PLACEHOLDER_PHONE = "0000000000"
PLACEHOLDER_GENDER = "Other"

TEMPLATE_HEADERS = [
    "Admission No", "First Name", "Last Name", "Class", "Section",
    "Roll No", "Date of Birth", "Gender", "Parent Phone", "Hosteler", "Transport",
]


def normalise_phone(raw) -> str:
    digits = re.sub(r"\D", "", str(raw) if raw is not None else "")
    if len(digits) >= 10:
        return digits[-10:]
    return PLACEHOLDER_PHONE


def split_name(full: str):
    parts = full.strip().split(None, 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return parts[0], ""


def parse_dob(cell) -> str:
    if cell is None:
        return PLACEHOLDER_DOB
    if isinstance(cell, (datetime, date)):
        d = cell.date() if isinstance(cell, datetime) else cell
        return d.strftime("%Y-%m-%d")
    s = str(cell).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return PLACEHOLDER_DOB


def find_col(hdr: list[str], *candidates: str) -> int | None:
    for c in candidates:
        for i, h in enumerate(hdr):
            if c.lower() in h.lower():
                return i
    return None


def read_source(path: Path, prefix: str, adm_year: str, class_name: str, section_name: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None and str(c).strip() for c in r)]
    wb.close()

    hdr = [str(c).strip() if c else "" for c in rows[0]]
    ci_roll  = find_col(hdr, "roll")
    ci_name  = find_col(hdr, "student", "name")
    ci_phone = find_col(hdr, "mobile", "phone")
    ci_dob   = find_col(hdr, "birth", "dob")

    records = []
    for row in rows[1:]:
        roll_raw = row[ci_roll] if ci_roll is not None else None
        if roll_raw is None or str(roll_raw).strip() == "":
            continue
        roll = int(str(roll_raw).strip().split(".")[0])
        name_raw = str(row[ci_name]).strip() if ci_name is not None and row[ci_name] else ""
        if not name_raw:
            continue

        first, last = split_name(name_raw)
        phone = normalise_phone(row[ci_phone]) if ci_phone is not None else PLACEHOLDER_PHONE
        dob   = parse_dob(row[ci_dob]) if ci_dob is not None else PLACEHOLDER_DOB
        adm_no = f"{prefix}-{adm_year}-{roll:03d}"

        records.append({
            "Admission No":  adm_no,
            "First Name":    first,
            "Last Name":     last,
            "Class":         class_name,
            "Section":       section_name,
            "Roll No":       str(roll),
            "Date of Birth": dob,
            "Gender":        PLACEHOLDER_GENDER,
            "Parent Phone":  phone,
            "Hosteler":      "No",
            "Transport":     "No",
        })
    return records


def build_template_xlsx(records: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(TEMPLATE_HEADERS)
    for rec in records:
        ws.append([rec[h] for h in TEMPLATE_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def main():
    database_url = os.environ["DATABASE_URL"]
    conn = await asyncpg.connect(database_url)

    tenant = await conn.fetchrow("SELECT id FROM tenants WHERE slug = $1", TENANT_SLUG)
    if not tenant:
        print(f"ERROR: tenant '{TENANT_SLUG}' not found"); await conn.close(); return

    ay = await conn.fetchrow(
        "SELECT id FROM academic_years WHERE tenant_id = $1 AND is_current = TRUE", tenant["id"]
    )
    if not ay:
        print("ERROR: no current academic year found"); await conn.close(); return

    print(f"Tenant: {TENANT_SLUG}  ({tenant['id']})")
    print(f"Academic year: {ay['id']}")

    all_records: list[dict] = []
    for rel_path, prefix, adm_year, class_name, section_name in SOURCES:
        src = BASE / rel_path
        if not src.exists():
            print(f"SKIP (not found): {src}"); continue
        recs = read_source(src, prefix, adm_year, class_name, section_name)
        print(f"  {prefix}-{adm_year} {class_name} {section_name}: {len(recs)} students  "
              f"(adm {prefix}-{adm_year}-001 … {prefix}-{adm_year}-{recs[-1]['Roll No'].zfill(3) if recs else '?'})")
        all_records.extend(recs)

    print(f"\nTotal rows to import: {len(all_records)}")
    if not all_records:
        print("Nothing to import."); await conn.close(); return

    xlsx_bytes = build_template_xlsx(all_records)

    result = await import_students(conn, tenant["id"], ay["id"], xlsx_bytes)
    await conn.close()

    print(f"\n=== Import result ===")
    print(f"  created : {result['created']}")
    print(f"  updated : {result['updated']}")
    if result["errors"]:
        print(f"  errors  : {len(result['errors'])}")
        for e in result["errors"]:
            print(f"    {e}")
    else:
        print(f"  errors  : 0")

    if result["errors"]:
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
