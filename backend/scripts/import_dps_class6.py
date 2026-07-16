"""
Import the real Class 6 roster for Daffodils Public School (DPS).

One-off follow-up to reset_and_import_dps.py: at the time of that run, the
source file's 19 Class-6 rows were placeholder roll numbers only (no name/
DOB/gender/phone), so Class 6 was created as an empty shell. The owner has
since supplied the real roster in the same source file. This script imports
just those rows into the existing tenant + current academic year — it does
NOT reset/delete anything, so it's safe to run standalone (upserts on
admission number, matching services.student.import_students).

Admission-number convention (matches reset_and_import_dps.py): DPS6-2026-{roll:03d}
"""

import asyncio
import io
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import asyncpg
import openpyxl
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent))
from services.student import import_students

TENANT_SLUG = "daffodilspublicschool"
SOURCE_XLSX = Path(__file__).parent.parent.parent / "School_docs" / "Daffodils" / "STUDENTS  INFO. $.xlsx"

ADM_YEAR = "2026"
CLASS_LABEL = "Class 6"
PREFIX = "DPS6"

PLACEHOLDER_DOB = "2000-01-01"
PLACEHOLDER_PHONE = "0000000000"

TEMPLATE_HEADERS = [
    "Admission No", "First Name", "Last Name", "Class", "Section",
    "Roll No", "Date of Birth", "Gender", "Parent Phone", "Hosteler", "Transport",
]


def normalise_phone(raw) -> str:
    digits = re.sub(r"\D", "", str(raw) if raw is not None else "")
    if len(digits) >= 10:
        return digits[-10:]
    return PLACEHOLDER_PHONE


def parse_dob(cell) -> str:
    if cell is None:
        return PLACEHOLDER_DOB
    if isinstance(cell, (datetime, date)):
        d = cell.date() if isinstance(cell, datetime) else cell
        return d.strftime("%Y-%m-%d")
    s = str(cell).strip().upper().replace("O", "0")
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[.\-/]+", ".", s)
    try:
        return datetime.strptime(s, "%d.%m.%Y").strftime("%Y-%m-%d")
    except ValueError:
        return PLACEHOLDER_DOB


def read_source() -> list[dict]:
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None and str(c).strip() for c in r)]
    wb.close()

    records = []
    skipped_no_data = 0
    for row in rows[1:]:
        if row[3] != 6:
            continue

        first = str(row[1]).strip() if row[1] else ""
        last = str(row[2]).strip() if row[2] else ""
        if not first:
            skipped_no_data += 1
            continue

        roll = int(str(row[5]).strip())
        gender_raw = str(row[6]).strip().upper() if row[6] else ""
        gender = {"M": "Male", "F": "Female", "MALE": "Male", "FEMALE": "Female"}.get(gender_raw, "Other")
        dob = parse_dob(row[7])
        phone = normalise_phone(row[10])
        transport = "Yes" if row[11] and str(row[11]).strip().upper() == "YES" else "No"
        hosteler = "Yes" if row[12] and str(row[12]).strip().upper() == "YES" else "No"

        records.append({
            "Admission No":  f"{PREFIX}-{ADM_YEAR}-{roll:03d}",
            "First Name":    first,
            "Last Name":     last,
            "Class":         CLASS_LABEL,
            "Section":       "A",
            "Roll No":       str(roll),
            "Date of Birth": dob,
            "Gender":        gender,
            "Parent Phone":  phone,
            "Hosteler":      hosteler,
            "Transport":     transport,
        })

    print(f"Parsed {len(records)} Class 6 rows (skipped {skipped_no_data} blank-name rows)")
    return records


def build_template_xlsx(records: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(TEMPLATE_HEADERS)
    for rec in records:
        ws.append([rec[h] for h in TEMPLATE_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def main():
    database_url = os.environ["DATABASE_URL"]
    conn = await asyncpg.connect(database_url)

    tenant = await conn.fetchrow("SELECT id FROM tenants WHERE slug = $1", TENANT_SLUG)
    if not tenant:
        print(f"ERROR: tenant '{TENANT_SLUG}' not found"); await conn.close(); return
    tenant_id = tenant["id"]

    ay = await conn.fetchrow(
        "SELECT id FROM academic_years WHERE tenant_id = $1 AND is_current = TRUE", tenant_id
    )
    if not ay:
        print("ERROR: no current academic year found"); await conn.close(); return

    records = read_source()
    if not records:
        print("Nothing to import."); await conn.close(); return

    xlsx_bytes = build_template_xlsx(records)
    result = await import_students(conn, tenant_id, ay["id"], xlsx_bytes)
    await conn.close()

    print("\n=== Import result ===")
    print(f"  created : {result['created']}")
    print(f"  updated : {result['updated']}")
    if result["errors"]:
        print(f"  errors  : {len(result['errors'])}")
        for e in result["errors"]:
            print(f"    {e}")
    else:
        print("  errors  : 0")

    if result["errors"]:
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
