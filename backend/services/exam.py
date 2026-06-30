import io
import uuid
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

import asyncpg

from core.events import emit
from models.exam import (
    BulkMarkRequest,
    ComponentMarksGrid,
    ConfigureComponentsRequest,
    ConsolidatedResult,
    ConsolidatedSubject,
    ExamComponentResponse,
    ExamSubjectCreate,
    ExamSubjectResponse,
    ExamTermCreate,
    ExamTermResponse,
    MarkEntryResponse,
    MarksConfigCreate,
    MarksConfigResponse,
    SaveComponentMarksRequest,
    StudentComponentRow,
    StudentTermResult,
    SubjectResult,
    TermResultSheet,
    TermStatusRequest,
)


class ExamError(Exception):
    pass


# ── Grade Calculation (CBSE scale) ────────────────────────────────────────────

def grade_from_pct(pct: Optional[float]) -> str:
    if pct is None:
        return "AB"
    if pct >= 91: return "A1"
    if pct >= 81: return "A2"
    if pct >= 71: return "B1"
    if pct >= 61: return "B2"
    if pct >= 51: return "C1"
    if pct >= 41: return "C2"
    if pct >= 33: return "D"
    return "E"


# ── Exam Subjects ─────────────────────────────────────────────────────────────

async def create_subject(
    conn: asyncpg.Connection, tenant_id: uuid.UUID, data: ExamSubjectCreate
) -> ExamSubjectResponse:
    try:
        row = await conn.fetchrow(
            """
            INSERT INTO exam_subjects
                (tenant_id, academic_year_id, class_id, name, subject_code, sort_order)
            VALUES ($1,$2,$3,$4,$5,$6)
            RETURNING *
            """,
            tenant_id, data.academic_year_id, data.class_id,
            data.name, data.subject_code, data.sort_order,
        )
    except asyncpg.UniqueViolationError:
        raise ExamError("A subject with this name already exists for this class and year")
    return ExamSubjectResponse(**dict(row))


async def list_subjects(
    conn: asyncpg.Connection,
    tenant_id: uuid.UUID,
    academic_year_id: Optional[uuid.UUID] = None,
    class_id: Optional[uuid.UUID] = None,
) -> list[ExamSubjectResponse]:
    rows = await conn.fetch(
        """
        SELECT * FROM exam_subjects
        WHERE tenant_id = $1
          AND ($2::uuid IS NULL OR academic_year_id = $2::uuid)
          AND ($3::uuid IS NULL OR class_id = $3::uuid)
          AND is_active = TRUE
        ORDER BY sort_order, name
        """,
        tenant_id, academic_year_id, class_id,
    )
    return [ExamSubjectResponse(**dict(r)) for r in rows]


async def deactivate_subject(
    conn: asyncpg.Connection, tenant_id: uuid.UUID, subject_id: uuid.UUID
) -> bool:
    """Soft-remove a subject (is_active=FALSE). Preserves any marks/timetable/homework
    that reference it — list_subjects already filters on is_active, so it disappears
    from the active picker. Returns False if the subject doesn't exist for this tenant."""
    row = await conn.fetchrow(
        """
        UPDATE exam_subjects SET is_active = FALSE
        WHERE id = $1 AND tenant_id = $2 AND is_active = TRUE
        RETURNING id
        """,
        subject_id, tenant_id,
    )
    return row is not None


# ── Exam Terms ────────────────────────────────────────────────────────────────

async def create_term(
    conn: asyncpg.Connection, tenant_id: uuid.UUID, data: ExamTermCreate
) -> ExamTermResponse:
    try:
        row = await conn.fetchrow(
            """
            INSERT INTO exam_terms
                (tenant_id, academic_year_id, name, term_type, start_date, end_date, sort_order)
            VALUES ($1,$2,$3,$4,$5,$6,$7)
            RETURNING *
            """,
            tenant_id, data.academic_year_id, data.name, data.term_type,
            data.start_date, data.end_date, data.sort_order,
        )
    except asyncpg.UniqueViolationError:
        raise ExamError("An exam term with this name already exists for this academic year")
    return ExamTermResponse(**dict(row))


async def list_terms(
    conn: asyncpg.Connection,
    tenant_id: uuid.UUID,
    academic_year_id: Optional[uuid.UUID] = None,
) -> list[ExamTermResponse]:
    rows = await conn.fetch(
        """
        SELECT * FROM exam_terms
        WHERE tenant_id = $1 AND ($2::uuid IS NULL OR academic_year_id = $2::uuid)
        ORDER BY sort_order, start_date NULLS LAST, name
        """,
        tenant_id, academic_year_id,
    )
    return [ExamTermResponse(**dict(r)) for r in rows]


async def publish_term(
    conn: asyncpg.Connection, tenant_id: uuid.UUID, term_id: uuid.UUID, publish: bool
) -> Optional[ExamTermResponse]:
    async with conn.transaction():
        prev = await conn.fetchval(
            "SELECT is_published FROM exam_terms WHERE id = $1 AND tenant_id = $2 FOR UPDATE",
            term_id, tenant_id,
        )
        if prev is None:
            return None
        row = await conn.fetchrow(
            "UPDATE exam_terms SET is_published = $1 WHERE id = $2 AND tenant_id = $3 RETURNING *",
            publish, term_id, tenant_id,
        )
        # Emit only on the false→true transition so re-publishing is a no-op
        # for downstream consumers (worker fan-out to parents).
        if publish and not prev:
            await emit(conn, "EXAM_PUBLISHED", tenant_id, {
                "term_id": str(term_id),
                "academic_year_id": str(row["academic_year_id"]),
            })
    return ExamTermResponse(**dict(row)) if row else None


# Valid transitions: which statuses can move to which next status.
_ALLOWED_TRANSITIONS: dict[str, set[str]] = {
    "draft":      {"marks_open"},
    "marks_open": {"locked"},
    "locked":     {"published", "marks_open"},   # marks_open = reopen
    "published":  {"locked"},                    # locked = reopen (principal escape hatch)
}

_EVENT_FOR_TRANSITION: dict[tuple[str, str], str] = {
    ("draft",      "marks_open"):  "EXAM_MARKS_OPENED",
    ("marks_open", "locked"):      "EXAM_MARKS_LOCKED",
    ("locked",     "published"):   "EXAM_PUBLISHED",
    ("locked",     "marks_open"):  "EXAM_REOPENED",
    ("published",  "locked"):      "EXAM_REOPENED",
}


async def transition_term_status(
    conn: asyncpg.Connection,
    tenant_id: uuid.UUID,
    term_id: uuid.UUID,
    new_status: str,
) -> ExamTermResponse:
    valid_statuses = {"draft", "marks_open", "locked", "published"}
    if new_status not in valid_statuses:
        raise ExamError(f"Invalid status '{new_status}'. Must be one of: {', '.join(sorted(valid_statuses))}")

    async with conn.transaction():
        row = await conn.fetchrow(
            "SELECT * FROM exam_terms WHERE id = $1 AND tenant_id = $2 FOR UPDATE",
            term_id, tenant_id,
        )
        if not row:
            raise ExamError("Exam term not found")

        current = row["status"]
        if new_status == current:
            return ExamTermResponse(**dict(row))

        allowed = _ALLOWED_TRANSITIONS.get(current, set())
        if new_status not in allowed:
            raise ExamError(
                f"Cannot move from '{current}' to '{new_status}'. "
                f"Allowed next states: {', '.join(sorted(allowed)) or 'none'}"
            )

        is_published = new_status == "published"
        updated = await conn.fetchrow(
            "UPDATE exam_terms SET status = $1, is_published = $2 WHERE id = $3 AND tenant_id = $4 RETURNING *",
            new_status, is_published, term_id, tenant_id,
        )

        event_type = _EVENT_FOR_TRANSITION.get((current, new_status))
        if event_type:
            await emit(conn, event_type, tenant_id, {
                "term_id": str(term_id),
                "academic_year_id": str(row["academic_year_id"]),
                "from_status": current,
                "to_status": new_status,
            })

    return ExamTermResponse(**dict(updated))


# ── Marks Config ──────────────────────────────────────────────────────────────

async def upsert_marks_config(
    conn: asyncpg.Connection, tenant_id: uuid.UUID, data: MarksConfigCreate
) -> MarksConfigResponse:
    row = await conn.fetchrow(
        """
        INSERT INTO exam_marks_config
            (tenant_id, exam_term_id, exam_subject_id, max_marks, passing_marks, weightage)
        VALUES ($1,$2,$3,$4,$5,$6)
        ON CONFLICT (tenant_id, exam_term_id, exam_subject_id)
        DO UPDATE SET max_marks = EXCLUDED.max_marks,
                      passing_marks = EXCLUDED.passing_marks,
                      weightage = EXCLUDED.weightage
        RETURNING *
        """,
        tenant_id, data.exam_term_id, data.exam_subject_id,
        data.max_marks, data.passing_marks, data.weightage,
    )
    subj = await conn.fetchval("SELECT name FROM exam_subjects WHERE id = $1", data.exam_subject_id)
    d = dict(row)
    d["subject_name"] = subj
    return MarksConfigResponse(**d)


async def list_marks_config(
    conn: asyncpg.Connection, tenant_id: uuid.UUID, exam_term_id: uuid.UUID
) -> list[MarksConfigResponse]:
    rows = await conn.fetch(
        """
        SELECT emc.*, es.name AS subject_name
        FROM exam_marks_config emc
        JOIN exam_subjects es ON es.id = emc.exam_subject_id
        WHERE emc.tenant_id = $1 AND emc.exam_term_id = $2
        ORDER BY es.sort_order, es.name
        """,
        tenant_id, exam_term_id,
    )
    return [MarksConfigResponse(**dict(r)) for r in rows]


# ── Mark Entries ──────────────────────────────────────────────────────────────

async def _assert_marks_open(conn: asyncpg.Connection, tenant_id: uuid.UUID, term_id: uuid.UUID) -> None:
    status = await conn.fetchval(
        "SELECT status FROM exam_terms WHERE id = $1 AND tenant_id = $2", term_id, tenant_id
    )
    if status is None:
        raise ExamError("Exam term not found")
    if status != "marks_open":
        label = {"draft": "Draft", "locked": "Locked", "published": "Published"}.get(status, status)
        raise ExamError(f"Marks entry is closed — term is {label}. Ask the principal to open marks.")


async def save_marks(
    conn: asyncpg.Connection,
    tenant_id: uuid.UUID,
    user_id: uuid.UUID,
    req: BulkMarkRequest,
) -> dict:
    await _assert_marks_open(conn, tenant_id, req.exam_term_id)

    # Validate marks don't exceed configured max_marks
    scored = [e for e in req.entries if not e.is_absent and e.marks_obtained is not None]
    if scored:
        subject_ids = list({e.exam_subject_id for e in scored})
        cfg_rows = await conn.fetch(
            """SELECT exam_subject_id, max_marks FROM exam_marks_config
               WHERE tenant_id=$1 AND exam_term_id=$2 AND exam_subject_id=ANY($3::uuid[])""",
            tenant_id, req.exam_term_id, subject_ids,
        )
        max_map = {row["exam_subject_id"]: row["max_marks"] for row in cfg_rows}
        for e in scored:
            cap = max_map.get(e.exam_subject_id)
            if cap is not None and e.marks_obtained > cap:
                raise ExamError(
                    f"Marks {e.marks_obtained} exceed the configured maximum {cap}"
                )

    await conn.executemany(
        """
        INSERT INTO mark_entries
            (tenant_id, student_id, exam_term_id, exam_subject_id,
             marks_obtained, is_absent, remarks, entered_by)
        VALUES ($1,$2,$3,$4,$5,$6,$7,$8)
        ON CONFLICT (tenant_id, student_id, exam_term_id, exam_subject_id)
        DO UPDATE SET
            marks_obtained = EXCLUDED.marks_obtained,
            is_absent      = EXCLUDED.is_absent,
            remarks        = EXCLUDED.remarks,
            entered_by     = EXCLUDED.entered_by,
            updated_at     = NOW()
        """,
        [
            (tenant_id, e.student_id, req.exam_term_id, e.exam_subject_id,
             e.marks_obtained, e.is_absent, e.remarks, user_id)
            for e in req.entries
        ],
    )
    await emit(conn, "MARKS_ENTERED", tenant_id, {
        "exam_term_id": str(req.exam_term_id),
        "count": len(req.entries),
    })
    return {"saved": len(req.entries)}


async def get_marks_for_term_subject(
    conn: asyncpg.Connection,
    tenant_id: uuid.UUID,
    exam_term_id: uuid.UUID,
    exam_subject_id: uuid.UUID,
    class_id: uuid.UUID,
    section_id: uuid.UUID,
) -> list[MarkEntryResponse]:
    rows = await conn.fetch(
        """
        SELECT me.*
        FROM mark_entries me
        JOIN students s ON s.id = me.student_id
        WHERE me.tenant_id = $1 AND me.exam_term_id = $2
          AND me.exam_subject_id = $3
          AND s.class_id = $4 AND s.section_id = $5
        ORDER BY
            CASE WHEN s.roll_number ~ '^[0-9]+$'
                 THEN LPAD(s.roll_number, 10, '0') ELSE s.roll_number END
        """,
        tenant_id, exam_term_id, exam_subject_id, class_id, section_id,
    )
    return [MarkEntryResponse(**dict(r)) for r in rows]


# ── Exam Components ────────────────────────────────────────────────────────────

async def configure_components(
    conn: asyncpg.Connection, tenant_id: uuid.UUID, req: ConfigureComponentsRequest
) -> list[ExamComponentResponse]:
    """Define the components for a (term, subject). Each component carries a
    weightage; the subject total is computed as a weighted percentage out of 100:
    SUM(marks/max * weightage) / SUM(weightage) * 100.
    We mirror max_marks=100 into exam_marks_config so downstream results stay /100."""
    if not req.components:
        raise ExamError("At least one component is required")

    keep_names = [c.name for c in req.components]

    async with conn.transaction():
        await conn.execute(
            """
            DELETE FROM exam_components
            WHERE tenant_id = $1 AND exam_term_id = $2 AND exam_subject_id = $3
              AND name <> ALL($4::text[])
            """,
            tenant_id, req.exam_term_id, req.exam_subject_id, keep_names,
        )
        for c in req.components:
            await conn.execute(
                """
                INSERT INTO exam_components
                    (tenant_id, exam_term_id, exam_subject_id, name, max_marks, weightage, sort_order)
                VALUES ($1,$2,$3,$4,$5,$6,$7)
                ON CONFLICT (tenant_id, exam_term_id, exam_subject_id, name)
                DO UPDATE SET max_marks = EXCLUDED.max_marks,
                              weightage = EXCLUDED.weightage,
                              sort_order = EXCLUDED.sort_order
                """,
                tenant_id, req.exam_term_id, req.exam_subject_id,
                c.name, c.max_marks, c.weightage, c.sort_order,
            )
        # Subject total is now scored out of 100 (weighted %); passing = 33
        await conn.execute(
            """
            INSERT INTO exam_marks_config
                (tenant_id, exam_term_id, exam_subject_id, max_marks, passing_marks, weightage)
            VALUES ($1,$2,$3,100,33,100)
            ON CONFLICT (tenant_id, exam_term_id, exam_subject_id)
            DO UPDATE SET max_marks = 100, passing_marks = 33
            """,
            tenant_id, req.exam_term_id, req.exam_subject_id,
        )

    return await list_components(conn, tenant_id, req.exam_term_id, req.exam_subject_id)


async def list_components(
    conn: asyncpg.Connection,
    tenant_id: uuid.UUID,
    exam_term_id: uuid.UUID,
    exam_subject_id: uuid.UUID,
) -> list[ExamComponentResponse]:
    rows = await conn.fetch(
        """
        SELECT id, name, max_marks, weightage, sort_order FROM exam_components
        WHERE tenant_id = $1 AND exam_term_id = $2 AND exam_subject_id = $3
        ORDER BY sort_order, name
        """,
        tenant_id, exam_term_id, exam_subject_id,
    )
    return [ExamComponentResponse(**dict(r)) for r in rows]


async def save_component_marks(
    conn: asyncpg.Connection,
    tenant_id: uuid.UUID,
    user_id: uuid.UUID,
    req: SaveComponentMarksRequest,
) -> dict:
    if not req.entries:
        return {"saved": 0}

    await _assert_marks_open(conn, tenant_id, req.exam_term_id)

    # Validate component marks don't exceed each component's max_marks
    scored = [e for e in req.entries if not e.is_absent and e.marks_obtained is not None]
    if scored:
        comp_ids = list({e.exam_component_id for e in scored})
        comp_rows = await conn.fetch(
            "SELECT id, name, max_marks FROM exam_components WHERE tenant_id=$1 AND id=ANY($2::uuid[])",
            tenant_id, comp_ids,
        )
        comp_max = {row["id"]: (row["name"], row["max_marks"]) for row in comp_rows}
        for e in scored:
            info = comp_max.get(e.exam_component_id)
            if info is not None:
                cname, cap = info
                if e.marks_obtained > cap:
                    raise ExamError(
                        f"Marks {e.marks_obtained} for '{cname}' exceed the maximum {cap}"
                    )

    async with conn.transaction():
        await conn.executemany(
            """
            INSERT INTO exam_component_marks
                (tenant_id, student_id, exam_component_id, marks_obtained, is_absent, entered_by)
            VALUES ($1,$2,$3,$4,$5,$6)
            ON CONFLICT (tenant_id, student_id, exam_component_id)
            DO UPDATE SET marks_obtained = EXCLUDED.marks_obtained,
                          is_absent      = EXCLUDED.is_absent,
                          entered_by     = EXCLUDED.entered_by,
                          updated_at     = NOW()
            """,
            [
                (tenant_id, e.student_id, e.exam_component_id, e.marks_obtained, e.is_absent, user_id)
                for e in req.entries
            ],
        )

        # Roll up into mark_entries using weighted formula: out of 100.
        # SUM(marks/max_marks * weightage) / SUM(weightage) * 100
        affected = list({e.student_id for e in req.entries})
        await conn.execute(
            """
            INSERT INTO mark_entries
                (tenant_id, student_id, exam_term_id, exam_subject_id, marks_obtained, is_absent, entered_by)
            SELECT $1, ecm.student_id, $2, $3,
                   ROUND(
                     SUM(CASE WHEN ecm.is_absent THEN 0
                              ELSE COALESCE(ecm.marks_obtained, 0) / ec.max_marks * ec.weightage END)
                     / NULLIF(SUM(ec.weightage), 0) * 100
                   , 2),
                   BOOL_AND(ecm.is_absent),
                   $4
            FROM exam_component_marks ecm
            JOIN exam_components ec ON ec.id = ecm.exam_component_id
            WHERE ec.tenant_id = $1 AND ec.exam_term_id = $2 AND ec.exam_subject_id = $3
              AND ecm.student_id = ANY($5::uuid[])
            GROUP BY ecm.student_id
            ON CONFLICT (tenant_id, student_id, exam_term_id, exam_subject_id)
            DO UPDATE SET marks_obtained = EXCLUDED.marks_obtained,
                          is_absent      = EXCLUDED.is_absent,
                          entered_by     = EXCLUDED.entered_by,
                          updated_at     = NOW()
            """,
            tenant_id, req.exam_term_id, req.exam_subject_id, user_id, affected,
        )

    await emit(conn, "MARKS_ENTERED", tenant_id, {
        "exam_term_id": str(req.exam_term_id),
        "exam_subject_id": str(req.exam_subject_id),
        "count": len(req.entries),
    })
    return {"saved": len(req.entries)}


async def get_component_marks_grid(
    conn: asyncpg.Connection,
    tenant_id: uuid.UUID,
    exam_term_id: uuid.UUID,
    exam_subject_id: uuid.UUID,
    class_id: uuid.UUID,
    section_id: uuid.UUID,
) -> ComponentMarksGrid:
    components = await list_components(conn, tenant_id, exam_term_id, exam_subject_id)
    total_max = Decimal("100")

    students = await conn.fetch(
        """
        SELECT s.id, s.first_name || ' ' || s.last_name AS full_name, s.roll_number
        FROM students s
        WHERE s.tenant_id = $1 AND s.class_id = $2 AND s.section_id = $3 AND s.is_active = TRUE
        ORDER BY CASE WHEN s.roll_number ~ '^[0-9]+$' THEN LPAD(s.roll_number, 10, '0') ELSE s.roll_number END
        """,
        tenant_id, class_id, section_id,
    )

    comp_ids = [c.id for c in components]
    marks = await conn.fetch(
        """
        SELECT ecm.student_id, ecm.exam_component_id, ecm.marks_obtained, ecm.is_absent
        FROM exam_component_marks ecm
        WHERE ecm.tenant_id = $1 AND ecm.exam_component_id = ANY($2::uuid[])
        """,
        tenant_id, comp_ids,
    ) if comp_ids else []
    mark_map = {(m["student_id"], m["exam_component_id"]): m for m in marks}

    weight_total = sum((c.weightage for c in components), Decimal("0"))
    rows: list[StudentComponentRow] = []
    for stu in students:
        per: dict[str, Optional[float]] = {}
        weighted_sum = Decimal("0")
        any_present = False
        all_absent = len(components) > 0
        for c in components:
            m = mark_map.get((stu["id"], c.id))
            if m and not m["is_absent"] and m["marks_obtained"] is not None:
                per[str(c.id)] = float(m["marks_obtained"])
                weighted_sum += Decimal(str(m["marks_obtained"])) / c.max_marks * c.weightage
                any_present = True
                all_absent = False
            else:
                per[str(c.id)] = None
                if not (m and m["is_absent"]):
                    all_absent = False
        pct = float(round(weighted_sum / weight_total * 100, 2)) if (any_present and weight_total) else None
        rows.append(StudentComponentRow(
            student_id=stu["id"],
            roll_number=stu["roll_number"],
            student_name=stu["full_name"],
            marks=per,
            is_absent=all_absent,
            total=pct,
        ))

    return ComponentMarksGrid(
        exam_term_id=exam_term_id,
        exam_subject_id=exam_subject_id,
        components=components,
        total_max=total_max,
        students=rows,
    )


# ── Result Computation ────────────────────────────────────────────────────────

async def compute_term_results(
    conn: asyncpg.Connection,
    tenant_id: uuid.UUID,
    exam_term_id: uuid.UUID,
    class_id: uuid.UUID,
    section_id: uuid.UUID,
) -> TermResultSheet:
    term = await conn.fetchrow(
        "SELECT name FROM exam_terms WHERE id = $1 AND tenant_id = $2", exam_term_id, tenant_id
    )
    if not term:
        raise ExamError("Exam term not found")

    configs = await conn.fetch(
        """
        SELECT emc.*, es.name AS subject_name, es.sort_order
        FROM exam_marks_config emc
        JOIN exam_subjects es ON es.id = emc.exam_subject_id
        WHERE emc.tenant_id = $1 AND emc.exam_term_id = $2 AND es.class_id = $3
        ORDER BY es.sort_order, es.name
        """,
        tenant_id, exam_term_id, class_id,
    )

    students = await conn.fetch(
        """
        SELECT s.id, s.first_name || ' ' || s.last_name AS full_name, s.roll_number, s.admission_no
        FROM students s
        WHERE s.tenant_id = $1 AND s.class_id = $2 AND s.section_id = $3 AND s.is_active = TRUE
        ORDER BY CASE WHEN s.roll_number ~ '^[0-9]+$' THEN LPAD(s.roll_number, 10, '0') ELSE s.roll_number END
        """,
        tenant_id, class_id, section_id,
    )

    # Fetch all mark entries for this term × this class-section
    entries = await conn.fetch(
        """
        SELECT me.student_id, me.exam_subject_id, me.marks_obtained, me.is_absent
        FROM mark_entries me
        JOIN students s ON s.id = me.student_id
        WHERE me.tenant_id = $1 AND me.exam_term_id = $2
          AND s.class_id = $3 AND s.section_id = $4
        """,
        tenant_id, exam_term_id, class_id, section_id,
    )
    entry_map = {(r["student_id"], r["exam_subject_id"]): r for r in entries}

    results = []
    for stu in students:
        subjects: list[SubjectResult] = []
        total_obtained = Decimal("0")
        total_max = Decimal("0")
        passed = True

        for cfg in configs:
            entry = entry_map.get((stu["id"], cfg["exam_subject_id"]))
            absent = entry["is_absent"] if entry else False
            marks = None if (not entry or absent) else entry["marks_obtained"]

            pct: Optional[float] = None
            if marks is not None and cfg["max_marks"] > 0:
                pct = round(float(marks) / float(cfg["max_marks"]) * 100, 1)

            subj_pass = (not absent) and (marks is not None) and (marks >= cfg["passing_marks"])
            if not subj_pass:
                passed = False

            if marks is not None:
                total_obtained += Decimal(str(marks))
                total_max += Decimal(str(cfg["max_marks"]))

            subjects.append(SubjectResult(
                subject_id=cfg["exam_subject_id"],
                subject_name=cfg["subject_name"],
                max_marks=Decimal(str(cfg["max_marks"])),
                marks_obtained=Decimal(str(marks)) if marks is not None else None,
                is_absent=absent,
                percentage=pct,
                grade=grade_from_pct(pct),
                passed=subj_pass,
            ))

        overall_pct = round(float(total_obtained) / float(total_max) * 100, 1) if total_max > 0 else None
        results.append(StudentTermResult(
            student_id=stu["id"],
            roll_number=stu["roll_number"],
            admission_no=stu["admission_no"],
            student_name=stu["full_name"],
            subjects=subjects,
            total_obtained=total_obtained,
            total_max=total_max,
            percentage=overall_pct,
            grade=grade_from_pct(overall_pct),
            passed=passed,
        ))

    return TermResultSheet(
        exam_term_id=exam_term_id,
        exam_term_name=term["name"],
        class_id=class_id,
        section_id=section_id,
        results=results,
    )


async def compute_consolidated_results(
    conn: asyncpg.Connection,
    tenant_id: uuid.UUID,
    academic_year_id: uuid.UUID,
    class_id: uuid.UUID,
    section_id: uuid.UUID,
) -> list[ConsolidatedResult]:
    """Weighted consolidated result across all published terms for a class-section."""
    terms = await conn.fetch(
        """
        SELECT * FROM exam_terms
        WHERE tenant_id = $1 AND academic_year_id = $2 AND is_published = TRUE
        ORDER BY sort_order, start_date NULLS LAST
        """,
        tenant_id, academic_year_id,
    )
    subjects = await conn.fetch(
        "SELECT * FROM exam_subjects WHERE tenant_id = $1 AND academic_year_id = $2 AND class_id = $3 AND is_active = TRUE ORDER BY sort_order, name",
        tenant_id, academic_year_id, class_id,
    )
    students = await conn.fetch(
        """
        SELECT s.id, s.first_name || ' ' || s.last_name AS full_name, s.roll_number, s.admission_no
        FROM students s
        WHERE s.tenant_id = $1 AND s.class_id = $2 AND s.section_id = $3 AND s.is_active = TRUE
        ORDER BY CASE WHEN s.roll_number ~ '^[0-9]+$' THEN LPAD(s.roll_number, 10, '0') ELSE s.roll_number END
        """,
        tenant_id, class_id, section_id,
    )

    # Get all configs and entries
    term_ids = [t["id"] for t in terms]
    if not term_ids:
        return []

    placeholders = ", ".join(f"${i + 2}" for i in range(len(term_ids)))
    configs = await conn.fetch(
        f"""
        SELECT emc.*, es.name AS subject_name
        FROM exam_marks_config emc
        JOIN exam_subjects es ON es.id = emc.exam_subject_id
        WHERE emc.tenant_id = $1 AND emc.exam_term_id IN ({placeholders})
        """,
        tenant_id, *term_ids,
    )
    n = len(term_ids)
    entries = await conn.fetch(
        f"""
        SELECT me.student_id, me.exam_term_id, me.exam_subject_id, me.marks_obtained, me.is_absent
        FROM mark_entries me
        JOIN students s ON s.id = me.student_id
        WHERE me.tenant_id = $1 AND me.exam_term_id IN ({placeholders})
          AND s.class_id = ${n + 2}::uuid AND s.section_id = ${n + 3}::uuid
        """,
        tenant_id, *term_ids, class_id, section_id,
    )

    cfg_map = {(c["exam_term_id"], c["exam_subject_id"]): c for c in configs}
    entry_map = {(e["student_id"], e["exam_term_id"], e["exam_subject_id"]): e for e in entries}

    results = []
    for stu in students:
        subj_results: list[ConsolidatedSubject] = []
        all_passed = True

        for subj in subjects:
            term_details = []
            weighted_sum = Decimal("0")
            weight_total = Decimal("0")
            subj_passed = True

            for term in terms:
                cfg = cfg_map.get((term["id"], subj["id"]))
                if not cfg:
                    continue
                entry = entry_map.get((stu["id"], term["id"], subj["id"]))
                absent = entry["is_absent"] if entry else False
                marks = None if (not entry or absent) else entry["marks_obtained"]

                if marks is not None:
                    weighted_sum += Decimal(str(marks)) / Decimal(str(cfg["max_marks"])) * Decimal(str(cfg["weightage"]))
                    weight_total += Decimal(str(cfg["weightage"]))
                else:
                    subj_passed = False

                term_details.append({
                    "term_name": term["name"],
                    "marks": float(marks) if marks is not None else None,
                    "max_marks": float(cfg["max_marks"]),
                    "weightage": float(cfg["weightage"]),
                    "is_absent": absent,
                })

            wp = round(float(weighted_sum / weight_total * 100), 1) if weight_total > 0 else None
            if not subj_passed:
                all_passed = False

            subj_results.append(ConsolidatedSubject(
                subject_id=subj["id"],
                subject_name=subj["name"],
                terms=term_details,
                weighted_percentage=wp,
                grade=grade_from_pct(wp),
                passed=subj_passed,
            ))

        valid_pcts = [s.weighted_percentage for s in subj_results if s.weighted_percentage is not None]
        overall_pct = round(sum(valid_pcts) / len(valid_pcts), 1) if valid_pcts else None
        results.append(ConsolidatedResult(
            student_id=stu["id"],
            roll_number=stu["roll_number"],
            admission_no=stu["admission_no"],
            student_name=stu["full_name"],
            subjects=subj_results,
            overall_percentage=overall_pct,
            overall_grade=grade_from_pct(overall_pct),
            passed=all_passed,
        ))

    return results


# ── Bulk Setup Import ─────────────────────────────────────────────────────────

_VALID_TERM_TYPES = {"unit_test", "half_yearly", "annual", "practical", "project", "internal", "term"}


async def import_exam_setup(
    conn: asyncpg.Connection,
    tenant_id: uuid.UUID,
    academic_year_id: uuid.UUID,
    file_bytes: bytes,
) -> dict:
    """
    Two-sheet xlsx:
      Sheet 'Subjects': Class | Subject | Subject Code | Sort Order
      Sheet 'Terms':    Term Name | Term Type | Start Date | End Date | Max Marks | Passing Marks
    Terms sheet sets marks config for ALL subjects imported in the same file.
    """
    try:
        import openpyxl
    except ImportError:
        raise ExamError("openpyxl not installed on server")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    if "subjects" not in sheet_names_lower:
        raise ExamError("Workbook must contain a sheet named 'Subjects'")
    if "terms" not in sheet_names_lower:
        raise ExamError("Workbook must contain a sheet named 'Terms'")

    subj_ws  = wb[sheet_names_lower["subjects"]]
    terms_ws = wb[sheet_names_lower["terms"]]

    class_rows = await conn.fetch("SELECT id, name FROM classes WHERE tenant_id = $1", tenant_id)
    class_map = {r["name"].strip().lower(): r["id"] for r in class_rows}

    # ── Subjects ──────────────────────────────────────────────────────────────
    subj_rows = list(subj_ws.iter_rows(values_only=True))
    if len(subj_rows) < 2:
        raise ExamError("Subjects sheet must have a header row and at least one data row")

    subj_hdr = [str(h).strip().lower() if h is not None else "" for h in subj_rows[0]]
    missing = {"class", "subject"} - set(subj_hdr)
    if missing:
        raise ExamError(f"Subjects sheet missing columns: {missing}")

    si = {n: subj_hdr.index(n) for n in ("class", "subject")}
    code_col  = subj_hdr.index("subject code") if "subject code" in subj_hdr else None
    order_col = subj_hdr.index("sort order")   if "sort order"   in subj_hdr else None

    subjects_created = subjects_updated = 0
    subj_errors: list[str] = []
    imported_subject_ids: list = []

    for row_num, row in enumerate(subj_rows[1:], start=2):
        try:
            class_str = str(row[si["class"]]).strip()
            subject   = str(row[si["subject"]]).strip()
            code      = str(row[code_col]).strip() if code_col is not None and row[code_col] else None
            sort_order = int(row[order_col]) if order_col is not None and row[order_col] else 0

            class_id = class_map.get(class_str.lower())
            if not class_id:
                raise ValueError(f"Class '{class_str}' not found")

            result = await conn.fetchrow(
                """
                INSERT INTO exam_subjects
                    (tenant_id, academic_year_id, class_id, name, subject_code, sort_order)
                VALUES ($1,$2,$3,$4,$5,$6)
                ON CONFLICT (tenant_id, academic_year_id, class_id, name)
                DO UPDATE SET subject_code = EXCLUDED.subject_code, sort_order = EXCLUDED.sort_order
                RETURNING id, (xmax = 0) AS inserted
                """,
                tenant_id, academic_year_id, class_id, subject, code, sort_order,
            )
            imported_subject_ids.append(result["id"])
            if result["inserted"]:
                subjects_created += 1
            else:
                subjects_updated += 1

        except Exception as exc:
            subj_errors.append(f"Subjects row {row_num}: {exc}")

    # ── Terms ─────────────────────────────────────────────────────────────────
    term_rows = list(terms_ws.iter_rows(values_only=True))
    if len(term_rows) < 2:
        raise ExamError("Terms sheet must have a header row and at least one data row")

    term_hdr = [str(h).strip().lower() if h is not None else "" for h in term_rows[0]]
    missing = {"term name", "term type", "max marks", "passing marks"} - set(term_hdr)
    if missing:
        raise ExamError(f"Terms sheet missing columns: {missing}")

    ti = {n: term_hdr.index(n) for n in ("term name", "term type", "max marks", "passing marks")}
    sd_col = term_hdr.index("start date") if "start date" in term_hdr else None
    ed_col = term_hdr.index("end date")   if "end date"   in term_hdr else None

    terms_created = terms_updated = 0
    term_errors: list[str] = []

    def _parse_date(v) -> date | None:
        if v is None: return None
        if isinstance(v, datetime): return v.date()
        if isinstance(v, date): return v
        return datetime.strptime(str(v).strip(), "%Y-%m-%d").date()

    for row_num, row in enumerate(term_rows[1:], start=2):
        try:
            name       = str(row[ti["term name"]]).strip()
            term_type  = str(row[ti["term type"]]).strip().lower()
            max_marks  = Decimal(str(row[ti["max marks"]])).quantize(Decimal("0.01"))
            pass_marks = Decimal(str(row[ti["passing marks"]])).quantize(Decimal("0.01"))
            start_date = _parse_date(row[sd_col]) if sd_col is not None else None
            end_date   = _parse_date(row[ed_col]) if ed_col is not None else None

            if term_type not in _VALID_TERM_TYPES:
                raise ValueError(f"Invalid term type '{term_type}'. Valid: {', '.join(sorted(_VALID_TERM_TYPES))}")

            result = await conn.fetchrow(
                """
                INSERT INTO exam_terms
                    (tenant_id, academic_year_id, name, term_type,
                     start_date, end_date, sort_order)
                VALUES ($1,$2,$3,$4,$5,$6,$7)
                ON CONFLICT (tenant_id, academic_year_id, name)
                DO UPDATE SET
                    term_type  = EXCLUDED.term_type,
                    start_date = EXCLUDED.start_date,
                    end_date   = EXCLUDED.end_date
                RETURNING id, (xmax = 0) AS inserted
                """,
                tenant_id, academic_year_id, name, term_type,
                start_date, end_date, row_num - 1,
            )
            term_id = result["id"]
            if result["inserted"]:
                terms_created += 1
            else:
                terms_updated += 1

            for sid in imported_subject_ids:
                await conn.execute(
                    """
                    INSERT INTO exam_marks_config
                        (tenant_id, exam_term_id, exam_subject_id, max_marks, passing_marks)
                    VALUES ($1,$2,$3,$4,$5)
                    ON CONFLICT (tenant_id, exam_term_id, exam_subject_id)
                    DO UPDATE SET max_marks = EXCLUDED.max_marks, passing_marks = EXCLUDED.passing_marks
                    """,
                    tenant_id, term_id, sid, max_marks, pass_marks,
                )

        except Exception as exc:
            term_errors.append(f"Terms row {row_num}: {exc}")

    return {
        "subjects_created": subjects_created,
        "subjects_updated": subjects_updated,
        "terms_created": terms_created,
        "terms_updated": terms_updated,
        "errors": subj_errors + term_errors,
    }
